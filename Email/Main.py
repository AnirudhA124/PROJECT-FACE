import os
from openpyxl import Workbook,load_workbook
from Email import send_email
import cv2

path='E:\FaceRecognition\Images'
images=[]
ClassNames=[]
myList=os.listdir(path)
#print(myList)

for cl in myList:
    img=cv2.imread(f'{path}/{cl}')
    images.append(img)
    ClassNames.append(os.path.splitext(cl)[0])
#print(ClassNames)

wb= load_workbook('E:\FaceRecognition\ATTENDANCE .xlsx')
ws= wb.active

perecentagecolumn= 'AK'
names= 'A'
emails= 'C'
people=len(ClassNames)+1
i='2'

while int(i)<people+1:
    if int(ws[perecentagecolumn+i].value)<75:
        send_email(
          subject="Attendance Reminder",
          name= str(ws[names+i].value),
          receiver_email=str(ws[emails+i].value),
          percentage=str(round(ws[perecentagecolumn+i].value,2))    
        )
    i=str(int(i)+1)
